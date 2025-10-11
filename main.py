import os
import json
import yaml
import zipfile
import pdfplumber
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from pymongo import MongoClient

# 🔹 Подключение к MongoDB
client = MongoClient("mongodb://ds_user:StrongPassword123@185.22.67.9:27017/yoyoflot?authSource=yoyoflot")
db = client["yoyoflot"]

DATA_DIR = "Airlines"

# ==============================
# Парсеры разных форматов
# ==============================

def load_csv(path, collection_name):
    df = pd.read_csv(path)
    db[collection_name].insert_many(df.to_dict(orient="records"))
    print(f"[+] CSV loaded: {collection_name} ({len(df)} records)")

def load_tab(path, collection_name):
    df = pd.read_csv(path, sep="\t")
    db[collection_name].insert_many(df.to_dict(orient="records"))
    print(f"[+] TAB loaded: {collection_name} ({len(df)} records)")

def load_json(path, collection_name):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
        if isinstance(data, dict):
            data = [data]
    db[collection_name].insert_many(data)
    print(f"[+] JSON loaded: {collection_name} ({len(data)} records)")

def load_yaml(path, collection_name):
    with open(path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
        if isinstance(data, dict):
            data = [data]
    db[collection_name].insert_many(data)
    print(f"[+] YAML loaded: {collection_name} ({len(data)} records)")

def load_xml(path, collection_name):
    tree = ET.parse(path)
    root = tree.getroot()
    records = []
    for item in root:
        record = {child.tag: child.text for child in item}
        records.append(record)
    db[collection_name].insert_many(records)
    print(f"[+] XML loaded: {collection_name} ({len(records)} records)")

def load_pdf(path, collection_name):
    text = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"
    db[collection_name].insert_one({"content": text})
    print(f"[+] PDF loaded: {collection_name}")

# ==============================
# Парсинг boarding pass (xlsx внутри zip)
# ==============================

def get_cell_value(ws, row, col):
    return ws.cell(row=row+1, column=col+1).value  # openpyxl индекс с 1

def load_boarding_pass(zip_path, collection_name):
    extract_dir = os.path.join(DATA_DIR, "unzipped")
    os.makedirs(extract_dir, exist_ok=True)

    with zipfile.ZipFile(zip_path, "r") as z:
        z.extractall(extract_dir)

    for f in os.listdir(extract_dir):
        if f.endswith(".xlsx"):
            fpath = os.path.join(extract_dir, f)
            wb = load_workbook(fpath)
            ws = wb.active

            flight_info = {
                'Flight Number': get_cell_value(ws, 3, 0),   # A4
                'Full Name': get_cell_value(ws, 2, 1),       # B3
                'Departure': get_cell_value(ws, 3, 3),       # D5
                'Arrival': get_cell_value(ws, 3, 7),         # H5
                'Departure Code': get_cell_value(ws, 6, 3),  # D7
                'Arrival Code': get_cell_value(ws, 6, 7),    # H7
                'Gate': get_cell_value(ws, 5, 1),            # B6
                'Flight Date': get_cell_value(ws, 7, 0),     # A9
                'Airline': get_cell_value(ws, 8, 4),         # E9
                'Unknown Time': get_cell_value(ws, 8, 2),    # C9
                'PNR': get_cell_value(ws, 11, 1),            # B12
                'E-Ticket': get_cell_value(ws, 11, 4),       # E13
                'Seat': get_cell_value(ws, 9, 5),            # F10
                'Unknown Number': get_cell_value(ws, 0, 7),  # H1
                'Sequence': get_cell_value(ws, 3, 5)         # F4
            }

            db[collection_name].insert_one(flight_info)
            print(f"[+] Boarding pass parsed: {f}, inserted into {collection_name}")

# ==============================
# Основная функция
# ==============================

def main():
    for file in os.listdir(DATA_DIR):
        path = os.path.join(DATA_DIR, file)
        cname = os.path.splitext(file)[0].lower()

        if file.endswith(".csv"):
            load_csv(path, cname)
        elif file.endswith(".tab"):
            load_tab(path, cname)
        elif file.endswith(".json"):
            load_json(path, cname)
        elif file.endswith(".yaml") or file.endswith(".yml"):
            load_yaml(path, cname)
        elif file.endswith(".xml"):
            load_xml(path, cname)
        elif file.endswith(".pdf"):
            load_pdf(path, cname)
        elif file.endswith(".zip"):
            load_boarding_pass(path, "boarding_pass")

    print("\n✅ Import complete. Collections in DB:")
    print(db.list_collection_names())

if __name__ == "__main__":
    main()
